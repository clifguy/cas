"""XLSX source adapter: extracts structural digest from Excel workbooks.

Produces a lightweight projection with sheet names as headings, column
headers, dimensions, and configurable preview rows rendered as
pipe-delimited tables. Designed for discovery, not full-content search;
agents retrieve the source.xlsx file for programmatic access.

Computes SHA-256 of raw.xlsx bytes for content_hash.
"""

import hashlib
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from sage.source_adapters.base import HeadingNode, ProjectionResult, SourceAdapter

# Default sheet names that signal "the user never renamed it"
_DEFAULT_SHEET_NAMES = {"Sheet", "Sheet1"}

_DEFAULT_PREVIEW_ROWS = 5


class XlsxAdapter(SourceAdapter):
    # 0.2.0: chunks indexed with heading-context (heading_path embedded with
    # content, plus FTS index on heading_path).
    # 0.3.0: chunker emits one chunk per heading regardless of body content
    # (Word-Find equivalence for empty-content parents).
    VERSION = "0.3.0"
    EXTENSIONS = [".xlsx"]

    async def project(self, source_path: Path, config: dict | None = None) -> ProjectionResult:
        config = config or {}
        preview_rows = config.get("preview_rows", _DEFAULT_PREVIEW_ROWS)
        max_sheets = config.get("max_sheets", None)

        raw_bytes = source_path.read_bytes()
        content_hash = hashlib.sha256(raw_bytes).hexdigest()

        wb = load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheet_names_all = wb.sheetnames
            sheets_to_process = sheet_names_all[:max_sheets] if max_sheets else sheet_names_all

            headings: list[HeadingNode] = []
            text_parts: list[str] = []
            dimensions_meta: dict[str, dict[str, int]] = {}

            for sheet_name in sheets_to_process:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                # Compute dimensions from actual data
                num_rows = len(rows)
                num_cols = max((len(r) for r in rows), default=0) if rows else 0

                dimensions_meta[sheet_name] = {
                    "rows": num_rows,
                    "columns": num_cols,
                }

                content_lines: list[str] = []
                content_lines.append(f"{num_rows} rows x {num_cols} columns")
                content_lines.append("")

                if rows:
                    # First row as header
                    header_row = rows[0]
                    header_cells = [str(c) if c is not None else "" for c in header_row]
                    content_lines.append("| " + " | ".join(header_cells) + " |")
                    content_lines.append("| " + " | ".join("---" for _ in header_cells) + " |")

                    # Preview data rows
                    data_rows = rows[1 : 1 + preview_rows]
                    for row in data_rows:
                        cells = [str(c) if c is not None else "" for c in row]
                        content_lines.append("| " + " | ".join(cells) + " |")

                    omitted = num_rows - 1 - len(data_rows)
                    if omitted > 0:
                        content_lines.append(f"... ({omitted} more data rows)")

                content = "\n".join(content_lines)

                headings.append(
                    HeadingNode(
                        level=1,
                        text=sheet_name,
                        path=sheet_name,
                        content=content,
                    )
                )

                text_parts.append(f"# {sheet_name}\n\n{content}")

            full_text = "\n\n".join(text_parts)

            # Title: first sheet name, unless it's a default name
            first_sheet = sheet_names_all[0] if sheet_names_all else None
            if first_sheet and first_sheet not in _DEFAULT_SHEET_NAMES:
                title = first_sheet
            else:
                title = source_path.stem

            source_mtime = datetime.fromtimestamp(source_path.stat().st_mtime, tz=timezone.utc)

            metadata = {
                "source_modified_at": source_mtime.isoformat(),
                "sheet_names": list(sheet_names_all),
                "total_sheets": len(sheet_names_all),
                "dimensions": dimensions_meta,
            }

        finally:
            wb.close()

        return ProjectionResult(
            text=full_text,
            headings=headings,
            content_hash=content_hash,
            adapter_version=self.VERSION,
            title=title,
            metadata=metadata,
        )
